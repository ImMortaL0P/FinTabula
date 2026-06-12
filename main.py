import os
import io
import re
import csv
import shutil
import uuid
from typing import List, Dict, Any, Optional
from fastapi import FastAPI, UploadFile, File, Form, HTTPException, status
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from dotenv import load_dotenv

from parser import extract_tables_local, extract_tables_ai

# Load environment variables
load_dotenv()

app = FastAPI(
    title="Financial Statement Table Extractor",
    description="Extract financial statement tables into beautifully formatted Excel sheets.",
    version="1.0.0"
)

# CORS Middleware (useful if developing or calling from other ports)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Create Temp directory for PDF uploads
TEMP_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "temp")
os.makedirs(TEMP_DIR, exist_ok=True)

# Pydantic schemas for Export Request
class TableExport(BaseModel):
    name: str
    data: List[List[str]]

class ExportRequest(BaseModel):
    tables: List[TableExport]
    format: str = "xlsx"  # "xlsx" or "csv"

@app.post("/api/extract")
async def extract_tables(
    file: UploadFile = File(...),
    mode: str = Form("local"),  # "local" or "gemini"
    pages: Optional[str] = Form(None),
    api_key: Optional[str] = Form(None),
    admin_user: Optional[str] = Form(None),
    admin_pass: Optional[str] = Form(None)
):
    # Reload env dynamically to pick up any changes
    load_dotenv(override=True)
    
    # Validate file format
    if not file.filename.endswith(".pdf"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only PDF files are supported."
        )
        
    # Resolve API Key: prioritize request payload, fallback to env variable
    resolved_api_key = api_key or os.getenv("GEMINI_API_KEY")
        
    # If mode is gemini, authenticate admin and check API key
    if mode == "gemini":
        if not admin_user or not admin_pass:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Admin credentials are required to use Gemini AI mode."
            )
        if admin_user != "Admin" or admin_pass != "Kukku404#":
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Invalid Admin username or password."
            )
        if not resolved_api_key:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Gemini API Key is required for AI-powered extraction."
            )

    # Save PDF locally in temp folder
    temp_filename = f"{uuid.uuid4()}_{file.filename}"
    temp_filepath = os.path.join(TEMP_DIR, temp_filename)
    
    try:
        with open(temp_filepath, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
            
        # Extract tables
        if mode == "gemini":
            tables = extract_tables_ai(temp_filepath, resolved_api_key, pages)
        else:
            tables = extract_tables_local(temp_filepath, pages)
            
        return {
            "success": True,
            "filename": file.filename,
            "mode": mode,
            "tables": tables
        }
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Extraction failed: {str(e)}"
        )
        
    finally:
        # Clean up temporary file
        if os.path.exists(temp_filepath):
            try:
                os.remove(temp_filepath)
            except Exception as ex:
                print(f"Error removing temp file: {ex}")

@app.post("/api/export")
async def export_tables(request: ExportRequest):
    tables = request.tables
    if not tables:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No tables provided for export."
        )
        
    if request.format == "csv":
        # Export first table as CSV using built-in csv module
        first_table = tables[0]
        output = io.StringIO()
        csv_writer = csv.writer(output)
        for row in first_table.data:
            csv_writer.writerow(row)
        output.seek(0)
        
        # Clean up filename
        safe_filename = re.sub(r'[^\w\-_\. ]', '_', first_table.name)
        headers = {
            'Content-Disposition': f'attachment; filename="{safe_filename}.csv"',
            'Access-Control-Expose-Headers': 'Content-Disposition'
        }
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode('utf-8')),
            headers=headers,
            media_type='text/csv'
        )
        
    elif request.format == "xlsx":
        import openpyxl
        wb = openpyxl.Workbook()
        # Remove default active sheet
        wb.remove(wb.active)
        
        # Populate sheets using openpyxl directly
        for idx, table in enumerate(tables):
            # Ensure sheet name is unique, valid, and <= 31 chars
            sheet_name = table.name.strip()
            if not sheet_name:
                sheet_name = f"Table {idx + 1}"
            
            # Replace invalid Excel sheet chars with underscore
            sheet_name = re.sub(r'[\\/?*:\[\]]', '_', sheet_name)
            sheet_name = sheet_name[:30]  # Cap length
            
            ws = wb.create_sheet(title=sheet_name)
            for row in table.data:
                ws.append(row)
                
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Apply premium formatting using openpyxl
        formatted_output = format_excel_workbook(output)
        
        headers = {
            'Content-Disposition': 'attachment; filename="extracted_financial_tables.xlsx"',
            'Access-Control-Expose-Headers': 'Content-Disposition'
        }
        return StreamingResponse(
            formatted_output,
            headers=headers,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unsupported format: {request.format}"
        )

def format_excel_workbook(file_bytes_io: io.BytesIO) -> io.BytesIO:
    """
    Loads an excel workbook and applies professional financial formatting:
    - Clean white background, no cell grids (borders) by default.
    - Centered, merged titles at the top.
    - Centered date headers with thin black bottom borders.
    - Bold section headers.
    - Bold total/subtotal rows with top/bottom underlines (on numeric columns only).
    - Parse numeric values (including currency signs and parentheses) into actual numbers.
    - Apply Excel Accounting format dynamically.
    - Auto-adjust column widths skipping title rows to avoid giant margins.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import re
    
    wb = openpyxl.load_workbook(file_bytes_io)
    
    for ws in wb.worksheets:
        # Enable gridlines in Excel window, but don't draw explicit borders on regular cells
        ws.views.sheetView[0].showGridLines = True
        
        # 1. Identify Title Rows and Date Header Rows
        title_rows = set()
        date_rows = set()
        
        def is_date_val(val_str):
            val_str = val_str.lower().strip()
            if not val_str:
                return False
            # Check for 4-digit years like 2021, 2020
            if re.search(r'\b(20\d\d)\b', val_str):
                return True
            # Check for month names
            months = ["january", "february", "march", "april", "may", "june", 
                      "july", "august", "september", "october", "november", "december"]
            if any(m in val_str for m in months):
                return True
            return False

        # Scan first 8 rows to classify titles and date headers
        for r_idx in range(1, min(9, ws.max_row + 1)):
            row_vals = [str(ws.cell(row=r_idx, column=c_idx).value or "").strip() for c_idx in range(1, ws.max_column + 1)]
            
            # Robust title check: if there is only 1 non-empty cell in the first 4 rows, it's a page title or subtitle
            non_empty_vals = [val for val in row_vals if val != ""]
            if len(non_empty_vals) == 1 and r_idx <= 4:
                single_val = non_empty_vals[0]
                # Consolidate it to column 1 and clear other columns to merge perfectly
                for c_idx in range(1, ws.max_column + 1):
                    ws.cell(row=r_idx, column=c_idx).value = None
                ws.cell(row=r_idx, column=1).value = single_val
                title_rows.add(r_idx)
            elif any(is_date_val(val) for val in row_vals):
                date_rows.add(r_idx)

        max_title_row = max(title_rows) if title_rows else 0

        # 2. Iterate through and format all cells
        for row_idx in range(1, ws.max_row + 1):
            row_cells = [ws.cell(row=row_idx, column=c_idx) for c_idx in range(1, ws.max_column + 1)]
            row_vals = [str(cell.value or "").strip() for cell in row_cells]
            col1_val = row_vals[0]
            
            # FOOTNOTE ROWS
            combined_text = " ".join(p for p in row_vals if p)
            is_footnote = (
                "accompanying notes" in combined_text.lower() or 
                "integral part of" in combined_text.lower() or
                "see notes to" in combined_text.lower()
            )
            
            if is_footnote:
                joined = ""
                for part in row_vals:
                    part = part.strip()
                    if not part:
                        continue
                    if not joined:
                        joined = part
                    else:
                        if joined[-1].isalpha() and joined[-1].islower() and part[0].isalpha() and part[0].islower():
                            joined += part
                        else:
                            joined += " " + part
                
                # Clear all cells in the row
                for col_idx in range(1, ws.max_column + 1):
                    c = ws.cell(row=row_idx, column=col_idx)
                    c.value = None
                    c.border = Border()
                    c.fill = PatternFill(fill_type=None)
                
                first_cell = ws.cell(row=row_idx, column=1)
                first_cell.value = joined
                first_cell.font = Font(name="Segoe UI", size=9, italic=True)
                first_cell.alignment = Alignment(horizontal="center", vertical="center")
                
                try:
                    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=ws.max_column)
                except Exception:
                    pass
                continue
            
            # TITLE ROWS
            if row_idx in title_rows:
                try:
                    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=ws.max_column)
                except Exception:
                    pass
                
                first_cell = ws.cell(row=row_idx, column=1)
                if row_idx == 1:
                    first_cell.font = Font(name="Segoe UI", size=14, bold=True)
                elif row_idx == 2:
                    first_cell.font = Font(name="Segoe UI", size=12, bold=True)
                elif row_idx == 3:
                    first_cell.font = Font(name="Segoe UI", size=10, italic=True)
                else:
                    first_cell.font = Font(name="Segoe UI", size=10)
                    
                first_cell.alignment = Alignment(horizontal="center", vertical="center")
                first_cell.fill = PatternFill(fill_type=None)
                first_cell.border = Border()
                
                # Make sure rest of the columns in the merged range have clean borders/fills
                for col_idx in range(2, ws.max_column + 1):
                    c = ws.cell(row=row_idx, column=col_idx)
                    c.font = Font(name="Segoe UI", size=10)
                    c.fill = PatternFill(fill_type=None)
                    c.border = Border()
                continue
                
            # DATE HEADER ROWS
            if row_idx in date_rows:
                for col_idx, cell in enumerate(row_cells, 1):
                    cell.font = Font(name="Segoe UI", size=10, bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.fill = PatternFill(fill_type=None)
                    # Apply single line underline below the years/dates (columns 2 to N)
                    if col_idx > 1:
                        cell.border = Border(bottom=Side(style='thin', color='000000'))
                    else:
                        cell.border = Border()
                continue
                
            # SECTION HEADERS
            # (No numbers in columns 2 to N, but Column 1 has text, and it's after titles)
            is_section_header = False
            if row_idx > max_title_row and col1_val:
                cols_2_to_n_empty = all(str(ws.cell(row=row_idx, column=c).value or "").strip() == "" for c in range(2, ws.max_column + 1))
                if cols_2_to_n_empty:
                    is_section_header = True
                    
            if is_section_header:
                for cell in row_cells:
                    cell.font = Font(name="Segoe UI", size=10, bold=True)
                    cell.fill = PatternFill(fill_type=None)
                    cell.border = Border()
                row_cells[0].alignment = Alignment(horizontal="left", vertical="center")
                continue

            # TOTAL / SUBTOTAL ROWS OR REGULAR DATA ROWS
            desc_lower = col1_val.lower()
            is_total_row = (
                desc_lower.startswith("total") or 
                "total" in desc_lower or 
                desc_lower.startswith("net income") or 
                desc_lower.startswith("net loss") or
                desc_lower.startswith("operating income") or 
                desc_lower.startswith("operating loss") or
                desc_lower.startswith("income before") or 
                desc_lower.startswith("net cash")
            )
            
            is_final_total = is_total_row and any(kw in desc_lower for kw in [
                "total assets", 
                "liabilities and equity", 
                "stockholders' equity", 
                "liabilities and stockholders", 
                "net income", 
                "net loss", 
                "total liabilities and equity"
            ])
            
            row_font = Font(name="Segoe UI", size=10, bold=is_total_row)
            
            for col_idx, cell in enumerate(row_cells, 1):
                cell.font = row_font
                cell.fill = PatternFill(fill_type=None)
                
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.border = Border()
                else:
                    # Value columns: Parse numbers and set accounting format
                    val_str = str(cell.value or "").strip()
                    is_numeric = False
                    num_val = 0
                    has_dollar = False
                    has_percent = False
                    has_decimal = False
                    
                    if val_str:
                        # Clean currency symbols and commas
                        clean_val = val_str.replace(",", "").replace(" ", "")
                        has_dollar = "$" in clean_val
                        has_percent = "%" in clean_val
                        clean_val = clean_val.replace("$", "").replace("%", "")
                        
                        # Handle negative numbers wrapped in parentheses
                        is_negative = False
                        if clean_val.startswith("(") and clean_val.endswith(")"):
                            is_negative = True
                            clean_val = clean_val[1:-1].strip()
                            
                        # Handle dashes (representing zero)
                        if clean_val in ["—", "-", "–"]:
                            is_numeric = True
                            num_val = 0
                        else:
                            # Match regular integer/float
                            if re.match(r'^\d+(\.\d+)?$', clean_val):
                                try:
                                    if "." in clean_val:
                                        num_val = float(clean_val)
                                        has_decimal = True
                                    else:
                                        num_val = int(clean_val)
                                    if is_negative:
                                        num_val = -num_val
                                    is_numeric = True
                                except ValueError:
                                    pass
                                    
                    if is_numeric:
                        # Convert cell value to a python number
                        if has_percent:
                            cell.value = num_val / 100.0 if isinstance(num_val, (int, float)) else num_val
                            cell.number_format = '0.0%' if has_decimal else '0%'
                        else:
                            cell.value = num_val
                            # Format with accounting style (either with or without dollar signs)
                            if has_dollar:
                                if has_decimal:
                                    cell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
                                else:
                                    cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                            else:
                                if has_decimal:
                                    cell.number_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
                                else:
                                    cell.number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                        
                    # Borders on total row values
                    if is_total_row:
                        if is_final_total:
                            # Top thin border, bottom double border
                            cell.border = Border(
                                top=Side(style='thin', color='000000'),
                                bottom=Side(style='double', color='000000')
                            )
                        else:
                            # Top thin border, bottom thin border
                            cell.border = Border(
                                top=Side(style='thin', color='000000'),
                                bottom=Side(style='thin', color='000000')
                            )
                    else:
                        cell.border = Border()

        # 3. Dynamic Column Auto-fitting (skipping title rows)
        for col_idx in range(1, ws.max_column + 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            max_len = 0
            for r_idx in range(1, ws.max_row + 1):
                if r_idx in title_rows:
                    continue
                cell = ws.cell(row=r_idx, column=col_idx)
                val = str(cell.value or "").strip()
                if cell.number_format and val:
                    # Provide extra padding for formatted numbers
                    if '$' in cell.number_format:
                        max_len = max(max_len, len(val) + 6)
                    else:
                        max_len = max(max_len, len(val) + 4)
                else:
                    for line in val.split("\n"):
                        max_len = max(max_len, len(line))
                        
            if col_idx == 1:
                ws.column_dimensions[col_letter].width = max(max_len + 4, 30)
            else:
                ws.column_dimensions[col_letter].width = max(max_len + 6, 14)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

# API Endpoint to check if server has a default API key
@app.get("/api/config")
async def get_config():
    load_dotenv(override=True)
    return {"has_api_key": bool(os.getenv("GEMINI_API_KEY"))}

# Serve index.html as homepage
@app.get("/")
async def get_index():
    index_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static", "index.html")
    if os.path.exists(index_path):
        return FileResponse(index_path)
    raise HTTPException(status_code=404, detail="Frontend homepage not found.")

# Mount the static directory
app.mount("/static", StaticFiles(directory=os.path.join(os.path.dirname(os.path.abspath(__file__)), "static")), name="static")
